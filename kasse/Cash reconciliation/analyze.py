"""
Cash Reconciliation Analyzer — CNC & BuckBuck
Reads kasseopgørelse (cash register statements) and produces a full breakdown:
  - Daily/weekly/monthly revenue by payment channel
  - Cash denomination distribution
  - Discrepancy detection (morning vs evening cash)
  - Employee-level summaries
  - Trend analysis
"""

import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime

DIR = Path(__file__).parent

# -- denomination columns (DKK) mapped to their multiplier --
DENOM_COLS = {
    "kr. 0,50": 0.50, "kr. 1": 1, "kr. 2": 2, "kr. 5": 5,
    "kr. 10": 10, "kr. 20": 20, "kr. 50": 50, "kr. 100": 100,
    "kr. 200": 200, "kr. 500": 500, "kr. 1000": 1000,
}


def load_sheet(filepath, sheet="24-25"):
    """Load the main sheet with data_only=True to resolve formulas where cached."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet]

    # read header row to map column letters -> names
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[cell.column] = cell.value

    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        record = {}
        for cell in row:
            col_name = headers.get(cell.column)
            if col_name and cell.value is not None:
                record[col_name] = cell.value
        if record.get("Dato"):  # skip empty rows
            rows.append(record)

    df = pd.DataFrame(rows)
    return df


def clean_df(df, source_name):
    """Standardize columns across CNC and BuckBuck files."""
    df = df.copy()
    df["Source"] = source_name

    # normalize date
    df["Dato"] = pd.to_datetime(df["Dato"], errors="coerce")
    df = df.dropna(subset=["Dato"])

    # fill missing numeric cols with 0
    num_cols = list(DENOM_COLS.keys()) + [
        "Online", "Kasse (POS Kontant)", "Kasse (POS Kreditkort)", "KIOSK",
    ]
    for c in num_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
        else:
            df[c] = 0

    # employee name
    if "Navn" in df.columns:
        df["Navn"] = df["Navn"].astype(str).replace("nan", "Unknown")
    else:
        df["Navn"] = "Unknown"

    # compute totals from raw denomination counts (actual cash in register)
    df["Cash_Total"] = sum(df[col] for col in DENOM_COLS.keys() if col in df.columns)

    # morning / evening holdings (may be formula-resolved or raw)
    for col in ["Kassebeholding morgen", "Kassebeholding aften Optalt beholding"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        else:
            df[col] = 0

    # daily total sales columns differ between files
    sale_col = "Dagens salg +" if "Dagens salg +" in df.columns else None
    if sale_col:
        df["Dagens_Salg"] = pd.to_numeric(df[sale_col], errors="coerce").fillna(0)
    else:
        # reconstruct: POS cash + POS credit + KIOSK + Online
        df["Dagens_Salg"] = (
            df["Kasse (POS Kontant)"]
            + df["Kasse (POS Kreditkort)"]
            + df["KIOSK"]
            + df["Online"]
        )

    df["Weekday"] = df["Dato"].dt.day_name()
    df["Week"] = df["Dato"].dt.isocalendar().week.astype(int)
    df["Month"] = df["Dato"].dt.to_period("M")

    return df


def print_section(title):
    print(f"\n{'='*60}")
    print(f"  {title}")
    print(f"{'='*60}")


def analyze(df, name):
    print_section(f"{name} — OVERVIEW")
    print(f"  Date range: {df['Dato'].min().date()} → {df['Dato'].max().date()}")
    print(f"  Total days recorded: {len(df)}")
    print(f"  Employees: {', '.join(df['Navn'].unique())}")

    # -- revenue summary --
    print_section(f"{name} — REVENUE SUMMARY")
    channels = {
        "POS Cash":   df["Kasse (POS Kontant)"].sum(),
        "POS Credit": df["Kasse (POS Kreditkort)"].sum(),
        "KIOSK":      df["KIOSK"].sum(),
        "Online":     df["Online"].sum(),
    }
    total_rev = sum(channels.values())
    for ch, val in channels.items():
        pct = (val / total_rev * 100) if total_rev else 0
        print(f"  {ch:15s}: {val:>12,.0f} DKK  ({pct:5.1f}%)")
    print(f"  {'TOTAL':15s}: {total_rev:>12,.0f} DKK")
    print(f"  Avg daily sale : {df['Dagens_Salg'].mean():>10,.0f} DKK")
    print(f"  Best day       : {df.loc[df['Dagens_Salg'].idxmax(), 'Dato'].date()}"
          f"  ({df['Dagens_Salg'].max():,.0f} DKK)")
    worst_idx = df.loc[df["Dagens_Salg"] > 0, "Dagens_Salg"].idxmin()
    if pd.notna(worst_idx):
        print(f"  Worst day      : {df.loc[worst_idx, 'Dato'].date()}"
              f"  ({df.loc[worst_idx, 'Dagens_Salg']:,.0f} DKK)")

    # -- day-of-week breakdown --
    print_section(f"{name} — DAY OF WEEK (avg daily sales)")
    dow = df.groupby("Weekday")["Dagens_Salg"].mean()
    day_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    for d in day_order:
        if d in dow.index:
            print(f"  {d:12s}: {dow[d]:>10,.0f} DKK")

    # -- monthly trend --
    print_section(f"{name} — MONTHLY TOTALS")
    monthly = df.groupby("Month").agg(
        Days=("Dato", "count"),
        Total_Sales=("Dagens_Salg", "sum"),
        Avg_Daily=("Dagens_Salg", "mean"),
        POS_Cash=("Kasse (POS Kontant)", "sum"),
        POS_Credit=("Kasse (POS Kreditkort)", "sum"),
        Kiosk=("KIOSK", "sum"),
        Online=("Online", "sum"),
    )
    for month, row in monthly.iterrows():
        print(f"  {str(month):8s} | {row['Days']:2.0f}d | "
              f"Total: {row['Total_Sales']:>10,.0f} | "
              f"Avg/day: {row['Avg_Daily']:>8,.0f} | "
              f"Cash: {row['POS_Cash']:>8,.0f} | "
              f"Card: {row['POS_Credit']:>8,.0f} | "
              f"Kiosk: {row['Kiosk']:>8,.0f} | "
              f"Online: {row['Online']:>7,.0f}")

    # -- cash denomination distribution --
    print_section(f"{name} — CASH DENOMINATION DISTRIBUTION (total across all days)")
    for denom, multiplier in DENOM_COLS.items():
        if denom in df.columns:
            raw = df[denom].sum()
            if raw > 0:
                count = raw / multiplier
                print(f"  {denom:12s}: {raw:>10,.0f} DKK  (~{count:>8,.0f} coins/bills)")

    # -- cash discrepancy (morning vs evening) --
    morning_col = "Kassebeholding morgen"
    evening_col = "Kassebeholding aften Optalt beholding"
    if df[morning_col].sum() > 0 or df[evening_col].sum() > 0:
        print_section(f"{name} — CASH DISCREPANCIES (evening − morning)")
        df_disc = df[["Dato", "Navn", morning_col, evening_col]].copy()
        df_disc["Diff"] = df_disc[evening_col] - df_disc[morning_col]
        # show days with notable discrepancies (> 500 DKK swing)
        big = df_disc[df_disc["Diff"].abs() > 500].sort_values("Diff")
        if len(big):
            print(f"  Days with >500 DKK discrepancy: {len(big)}")
            for _, r in big.iterrows():
                print(f"    {r['Dato'].date()} ({r['Navn']:>8s}): "
                      f"morning={r[morning_col]:>8,.0f}  evening={r[evening_col]:>8,.0f}  "
                      f"diff={r['Diff']:>+9,.0f}")
        else:
            print("  No major discrepancies (all within 500 DKK)")

    # -- employee breakdown --
    employees = df["Navn"].unique()
    if len(employees) > 1 or (len(employees) == 1 and employees[0] != "Unknown"):
        print_section(f"{name} — EMPLOYEE BREAKDOWN")
        emp = df.groupby("Navn").agg(
            Days=("Dato", "count"),
            Avg_Sale=("Dagens_Salg", "mean"),
            Total=("Dagens_Salg", "sum"),
        ).sort_values("Total", ascending=False)
        for name_emp, row in emp.iterrows():
            print(f"  {name_emp:12s}: {row['Days']:3.0f} days | "
                  f"avg {row['Avg_Sale']:>8,.0f}/day | "
                  f"total {row['Total']:>10,.0f} DKK")


def main():
    print("\n" + "▓" * 60)
    print("  KASSEOPGØRELSE ANALYSIS — CNC & BuckBuck")
    print("▓" * 60)

    # load both files
    cnc_path = DIR / "CNC Kasseopgørelse.xlsx"
    bb_path = DIR / "BuckBuck Kasseopgørelse.xlsx"

    frames = []
    for path, label in [(cnc_path, "CNC"), (bb_path, "BuckBuck")]:
        if path.exists():
            raw = load_sheet(path)
            df = clean_df(raw, label)
            frames.append((label, df))
            analyze(df, label)
        else:
            print(f"  ⚠ File not found: {path}")

    # -- combined comparison --
    if len(frames) == 2:
        print_section("COMBINED — CNC vs BuckBuck COMPARISON")
        for label, df in frames:
            total = df["Dagens_Salg"].sum()
            avg = df["Dagens_Salg"].mean()
            card_pct = (
                df["Kasse (POS Kreditkort)"].sum()
                / (df["Kasse (POS Kreditkort)"].sum() + df["Kasse (POS Kontant)"].sum() + 0.01)
                * 100
            )
            print(f"  {label:10s}: total {total:>12,.0f} DKK | "
                  f"avg/day {avg:>8,.0f} DKK | "
                  f"card share {card_pct:5.1f}% | "
                  f"days: {len(df)}")

    print(f"\n{'='*60}")
    print("  Analysis complete.")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
